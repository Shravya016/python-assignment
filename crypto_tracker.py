import requests
import pandas as pd
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os


COINGECKO_API_URL = "https://api.coingecko.com/api/v3"
EXCEL_FILE_PATH = "crypto_data_live.xlsx"
REFRESH_INTERVAL = 300  

def fetch_top_50_cryptos():
    try:
        endpoint = f"{COINGECKO_API_URL}/coins/markets"
        params = {
            "vs_currency": "usd",
            "order": "market_cap_desc",
            "per_page": 50,
            "page": 1,
            "sparkline": False,
            "price_change_percentage": "24h"
        }
        
        response = requests.get(endpoint, params=params)
        response.raise_for_status()  
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data from CoinGecko API: {e}")
        return None

def process_crypto_data(data):
    if not data:
        return None
    
    processed_data = []
    for coin in data:
        processed_data.append({
            "name": coin["name"],
            "symbol": coin["symbol"].upper(),
            "current_price": coin["current_price"],
            "market_cap": coin["market_cap"],
            "trading_volume_24h": coin["total_volume"],
            "price_change_24h": coin["price_change_percentage_24h"] if coin["price_change_percentage_24h"] else 0
        })
    
    return processed_data

def analyze_crypto_data(data):
    if not data:
        return None
    
    df = pd.DataFrame(data)
    
    analysis = {
        "top_5_by_market_cap": df.nlargest(5, "market_cap")[["name", "symbol", "market_cap"]].to_dict("records"),
        "average_price": df["current_price"].mean(),
        "highest_price_change": df.loc[df["price_change_24h"].idxmax()],
        "lowest_price_change": df.loc[df["price_change_24h"].idxmin()],
        "total_market_cap": df["market_cap"].sum(),
        "total_trading_volume": df["trading_volume_24h"].sum(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    return analysis

def update_excel(data, analysis):
    """Update the Excel file with cryptocurrency data and analysis."""
    if not data:
        return False
    
    df = pd.DataFrame(data)
    
    
    with pd.ExcelWriter(EXCEL_FILE_PATH, engine='openpyxl') as writer:
   
        df.to_excel(writer, sheet_name='Live Crypto Data', index=False)
        
        
        analysis_df = pd.DataFrame([{
            "Timestamp": analysis["timestamp"],
            "Average Price (USD)": analysis["average_price"],
            "Total Market Cap (USD)": analysis["total_market_cap"],
            "Total 24h Trading Volume (USD)": analysis["total_trading_volume"]
        }])
        analysis_df.to_excel(writer, sheet_name='Analysis', index=False, startrow=1)
        
        
        top5_df = pd.DataFrame(analysis["top_5_by_market_cap"])
        top5_df.to_excel(writer, sheet_name='Analysis', index=False, startrow=6)
        
        
        price_changes = pd.DataFrame([
            {
                "Type": "Highest 24h Price Change",
                "Name": analysis["highest_price_change"]["name"],
                "Symbol": analysis["highest_price_change"]["symbol"],
                "Price Change (%)": analysis["highest_price_change"]["price_change_24h"]
            },
            {
                "Type": "Lowest 24h Price Change",
                "Name": analysis["lowest_price_change"]["name"],
                "Symbol": analysis["lowest_price_change"]["symbol"],
                "Price Change (%)": analysis["lowest_price_change"]["price_change_24h"]
            }
        ])
        price_changes.to_excel(writer, sheet_name='Analysis', index=False, startrow=13)
        
        
        workbook = writer.book
        
        
        worksheet = writer.sheets['Live Crypto Data']
        for col in range(1, len(df.columns) + 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        
        analysis_sheet = writer.sheets['Analysis']
        analysis_sheet.merge_cells('A1:D1')
        analysis_sheet.cell(1, 1).value = "Cryptocurrency Market Analysis"
        analysis_sheet.cell(1, 1).font = Font(bold=True, size=14)
        
        analysis_sheet.merge_cells('A6:D6')
        analysis_sheet.cell(6, 1).value = "Top 5 Cryptocurrencies by Market Cap"
        analysis_sheet.cell(6, 1).font = Font(bold=True)
        
        analysis_sheet.merge_cells('A13:D13')
        analysis_sheet.cell(13, 1).value = "24-Hour Price Change Extremes"
        analysis_sheet.cell(13, 1).font = Font(bold=True)
    
    print(f"Excel file updated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    return True

def generate_analysis_report(data, analysis):
    """Generate a simple analysis report and save it as a text file."""
    if not data or not analysis:
        return False
    
    report_file = "crypto_analysis_report.txt"
    
    with open(report_file, "w") as f:
        f.write("CRYPTOCURRENCY MARKET ANALYSIS REPORT\n")
        f.write("====================================\n\n")
        f.write(f"Generated on: {analysis['timestamp']}\n\n")
        
        f.write("MARKET OVERVIEW\n")
        f.write("--------------\n")
        f.write(f"Total Market Cap of Top 50: ${analysis['total_market_cap']:,.2f}\n")
        f.write(f"Total 24h Trading Volume: ${analysis['total_trading_volume']:,.2f}\n")
        f.write(f"Average Price of Top 50: ${analysis['average_price']:,.2f}\n\n")
        
        f.write("TOP 5 CRYPTOCURRENCIES BY MARKET CAP\n")
        f.write("-----------------------------------\n")
        for i, coin in enumerate(analysis['top_5_by_market_cap'], 1):
            f.write(f"{i}. {coin['name']} ({coin['symbol']}): ${coin['market_cap']:,.2f}\n")
        f.write("\n")
        
        f.write("24-HOUR PRICE CHANGE EXTREMES\n")
        f.write("----------------------------\n")
        f.write(f"Highest: {analysis['highest_price_change']['name']} ({analysis['highest_price_change']['symbol']}): {analysis['highest_price_change']['price_change_24h']:.2f}%\n")
        f.write(f"Lowest: {analysis['lowest_price_change']['name']} ({analysis['lowest_price_change']['symbol']}): {analysis['lowest_price_change']['price_change_24h']:.2f}%\n\n")
        
        f.write("NOTE: For more detailed information and live updates, please refer to the Excel file.\n")
    
    print(f"Analysis report generated at {report_file}")
    return True

def main():
    """Main function to run the cryptocurrency tracker."""
    print("Cryptocurrency Live Tracker Started")
    print("==================================")
    print("Fetching and analyzing top 50 cryptocurrencies...")
    print("Press Ctrl+C to stop the program.")
    
    run_count = 0
    
    try:
        while True:
            run_count += 1
            print(f"\nUpdate #{run_count} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            
            crypto_data = fetch_top_50_cryptos()
            
            if crypto_data:
                
                processed_data = process_crypto_data(crypto_data)
                
               
                analysis_results = analyze_crypto_data(processed_data)
                
               
                update_excel(processed_data, analysis_results)
                
                
                if run_count == 1:
                    generate_analysis_report(processed_data, analysis_results)
                
                print(f"Data updated successfully. Next update in {REFRESH_INTERVAL // 60} minutes.")
            else:
                print("Failed to fetch data. Will retry in the next cycle.")
            
           
            time.sleep(REFRESH_INTERVAL)
            
    except KeyboardInterrupt:
        print("\nProgram terminated by user.")
    except Exception as e:
        print(f"\nAn error occurred: {e}")
    finally:
        print("\nCryptocurrency Live Tracker Stopped")

if __name__ == "__main__":
    main() 